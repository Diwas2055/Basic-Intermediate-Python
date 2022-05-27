# (A) INIT
# (A1) LOAD MODULES
import sqlite3, os
from sqlite3 import Error
from openpyxl import load_workbook

# (A2) SETTINGS
DBFILE = "users.db"

# (B) OPEN DATABASE & EXCEL FILE
conn = sqlite3.connect(DBFILE)
cursor = conn.cursor()
book = load_workbook("s3_users.xlsx")
sheet = book.active

# (C) IMPORT ROWS & COLUMNS
for row in range(1, sheet.max_row + 1):
  sql = "INSERT INTO `users` (`name`, `email`, `tel`) VALUES "
  sql += f"('{sheet.cell(row, 1).value}', '{sheet.cell(row, 2).value}', '{sheet.cell(row, 3).value}')"
  cursor.execute(sql)
  print(sql)
conn.commit()
conn.close()
